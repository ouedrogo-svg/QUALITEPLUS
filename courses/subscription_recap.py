"""Récapitulatif des abonnements approuvés par date réelle de validation."""

from __future__ import annotations

from collections import defaultdict
from decimal import Decimal
from io import BytesIO

from django.urls import reverse

from .models import (
    SubscriptionPlan,
    SubscriptionRequest,
    content_month_period_label,
    french_month_name,
)

RECAP_COLUMNS = (
    "Date de validation",
    "Nom",
    "Prénom",
    "Catégorie",
    "Mois d'abonnement",
    "Option d'abonnement",
    "Montant",
)


def _plan_option_label(plan: SubscriptionPlan) -> str:
    return plan.display_name()


def _content_month_label(req: SubscriptionRequest) -> str:
    """Mois de contenu couverts par la tranche."""
    periods = req.covered_periods()
    if not periods:
        return content_month_period_label(req.year, req.month)
    if len(periods) == 1:
        y, m = periods[0]
        return content_month_period_label(y, m)
    labels = [content_month_period_label(y, m) for y, m in periods]
    return ", ".join(labels)


def _format_amount(amount: Decimal) -> str:
    if amount == amount.to_integral_value():
        return f"{int(amount):,}".replace(",", " ")
    return f"{amount:,.2f}".replace(",", " ")


def _approved_requests_queryset(*, category_ids: set[int] | None = None):
    qs = (
        SubscriptionRequest.objects.filter(
            status=SubscriptionRequest.Status.APPROVED,
            decided_at__isnull=False,
        )
        .select_related("user", "category", "plan")
        .prefetch_related("plan__included_periods")
        .order_by("-decided_at")
    )
    if category_ids is not None:
        if not category_ids:
            return qs.none()
        qs = qs.filter(category_id__in=category_ids)
    return qs


def _row_from_request(req: SubscriptionRequest) -> dict:
    from django.utils import timezone
    sub_date = timezone.localdate(req.decided_at)
    plan = req.plan
    if plan:
        amount = plan.amount
        option_label = _plan_option_label(plan)
    else:
        amount = Decimal("0")
        option_label = "—"

    return {
        "date": sub_date,
        "date_display": sub_date.strftime("%d/%m/%Y"),
        "nom": (req.user.last_name or "").strip() or "—",
        "prenom": (req.user.first_name or "").strip() or "—",
        "categorie": req.category.name,
        "mois_abonnement": _content_month_label(req),
        "option": option_label,
        "montant": amount,
        "montant_display": _format_amount(amount),
    }


def subscription_recap_rows(
    *, for_month: tuple[int, int] | None = None, category_ids: set[int] | None = None
) -> list[dict]:
    rows = []
    # On filtre en Python pour garantir une cohérence parfaite avec le récap affiché à l'écran.
    # Le récap à l'écran (tree) utilise localdate() pour grouper, on doit faire pareil ici.
    for req in _approved_requests_queryset(category_ids=category_ids):
        row = _row_from_request(req)
        if for_month is not None:
            year, month = for_month
            if row["date"].year != year or row["date"].month != month:
                continue
        rows.append(row)

    rows.sort(key=lambda r: (r["date"], r["nom"].lower(), r["prenom"].lower()))
    return rows


def _month_total(rows: list[dict]) -> Decimal:
    return sum((r["montant"] for r in rows), Decimal("0"))


def build_subscription_recap_tree(
    *,
    month_export_url,
    category_ids: set[int] | None = None,
    include_rows: bool = True,
) -> list[dict]:
    """
    Récapitulatif par mois calendaire de validation (decided_at).
    ``month_export_url`` : callable ``(year, month) -> str`` pour le lien d’export du mois.
    """
    from django.utils import timezone
    months_data: dict[tuple[int, int], list[dict]] = defaultdict(list)

    for req in _approved_requests_queryset(category_ids=category_ids):
        row = _row_from_request(req)
        # On utilise le même localdate que dans _row_from_request pour la cohérence
        month_key = (row["date"].year, row["date"].month)
        months_data[month_key].append(row)

    tree = []
    for year, month in sorted(months_data.keys(), reverse=True):
        month_rows = sorted(
            months_data[(year, month)],
            key=lambda r: (r["date"], r["nom"].lower(), r["prenom"].lower()),
        )
        tree.append(
            {
                "year": year,
                "month": month,
                "label": f"{french_month_name(month)} {year}",
                "rows": month_rows if include_rows else [],
                "n_subscriptions": len(month_rows),
                "total_amount": _month_total(month_rows),
                "total_display": _format_amount(_month_total(month_rows)),
                "export_url": month_export_url(year, month),
            }
        )
    return tree


def subscription_recap_global_export_url() -> str:
    return reverse("admin:courses_subscriptionrequest_export_recap")


def subscription_recap_month_export_url(year: int, month: int) -> str:
    return reverse(
        "admin:courses_subscriptionrequest_export_recap_month",
        args=[f"{year:04d}-{month:02d}"],
    )


def build_admin_subscription_recap_tree(*, include_rows: bool = True) -> list[dict]:
    """Récap admin : liens d’export vers les vues ``admin``."""
    return build_subscription_recap_tree(
        month_export_url=subscription_recap_month_export_url,
        include_rows=include_rows,
    )


def formateur_subscription_recap_global_export_url() -> str:
    return reverse("courses:formateur_recap_export_all")


def formateur_subscription_recap_month_export_url(year: int, month: int) -> str:
    return reverse(
        "courses:formateur_recap_export_month",
        args=[f"{year:04d}-{month:02d}"],
    )


def build_formateur_subscription_recap_tree(user) -> list[dict]:
    """Récap pour l’espace formateur complet : URLs d’export du site public."""
    from .formateur_permissions import formateur_category_ids

    return build_subscription_recap_tree(
        month_export_url=formateur_subscription_recap_month_export_url,
        category_ids=formateur_category_ids(user),
    )


def formateur_contenu_subscription_recap_global_export_url() -> str:
    return reverse("courses:formateur_contenu_recap_export_all")


def formateur_contenu_subscription_recap_month_export_url(year: int, month: int) -> str:
    return reverse(
        "courses:formateur_contenu_recap_export_month",
        args=[f"{year:04d}-{month:02d}"],
    )


def build_formateur_contenu_subscription_recap_tree(user) -> list[dict]:
    """Récap pour l’espace formateur contenu (sans gestion des demandes)."""
    from .formateur_permissions import formateur_category_ids

    return build_subscription_recap_tree(
        month_export_url=formateur_contenu_subscription_recap_month_export_url,
        category_ids=formateur_category_ids(user, assigned_only=True),
    )


def subscription_recap_spreadsheet_rows(
    *,
    for_month: tuple[int, int] | None = None,
    category_ids: set[int] | None = None,
) -> list[tuple]:
    return [
        (
            row["date_display"],
            row["nom"],
            row["prenom"],
            row["categorie"],
            row["mois_abonnement"],
            row["option"],
            float(row["montant"]),
        )
        for row in subscription_recap_rows(
            for_month=for_month, category_ids=category_ids
        )
    ]


def subscription_recap_filename(
    *, for_month: tuple[int, int] | None = None, ext: str = "xlsx"
) -> str:
    from django.utils import timezone

    if for_month is not None:
        year, month = for_month
        return f"abonnements_{year:04d}-{month:02d}.{ext}"

    # Pour l'export global, on ajoute la date du jour
    today = timezone.localdate()
    return f"abonnements_global_{today.strftime('%Y-%m-%d')}.{ext}"


def _recap_title(*, for_month: tuple[int, int] | None = None) -> str:
    if for_month is not None:
        year, month = for_month
        return f"Abonnements approuvés — {french_month_name(month)} {year}"
    return "Récapitulatif des abonnements approuvés"


def build_subscription_recap_xlsx(
    *,
    for_month: tuple[int, int] | None = None,
    category_ids: set[int] | None = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    rows = subscription_recap_spreadsheet_rows(
        for_month=for_month, category_ids=category_ids
    )
    title = _recap_title(for_month=for_month)

    wb = Workbook()
    ws = wb.active
    ws.title = "Abonnements"

    ws.append([title])
    ws.append(
        ["Date réelle d’abonnement = date d’approbation par l’administrateur."]
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(RECAP_COLUMNS))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(RECAP_COLUMNS))
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])

    ws.append(list(RECAP_COLUMNS))
    header_row = ws.max_row
    n_cols = len(RECAP_COLUMNS)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2E8F0")
        cell.alignment = Alignment(horizontal="center")

    total = Decimal("0")
    for row_data in rows:
        ws.append(list(row_data))
        # row_data[-1] est le montant (float)
        total += Decimal(str(row_data[-1]))

    data_last_row = ws.max_row
    if rows:
        total_row = data_last_row + 2
        ws.cell(row=total_row, column=n_cols - 1, value="TOTAL GÉNÉRAL")
        ws.cell(row=total_row, column=n_cols - 1).font = Font(bold=True)
        ws.cell(row=total_row, column=n_cols - 1).alignment = Alignment(
            horizontal="right"
        )
        total_cell = ws.cell(row=total_row, column=n_cols, value=float(total))
        total_cell.font = Font(bold=True)
        total_cell.number_format = "#,##0"
        total_cell.fill = PatternFill("solid", fgColor="F1F5F9")
        total_cell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 14

    if rows:
        for row in ws.iter_rows(
            min_row=header_row + 1,
            max_row=data_last_row,
            min_col=n_cols,
            max_col=n_cols,
        ):
            amount_cell = row[0]
            if isinstance(amount_cell.value, (int, float)):
                amount_cell.number_format = "#,##0"
                amount_cell.alignment = Alignment(horizontal="right")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
